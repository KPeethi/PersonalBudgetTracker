"""
Excel data visualizer for the Expense Tracker application.
Generates various charts and visualizations from expense data in Excel files.
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import io
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Define constants for styling
TITLE_FONT = Font(name='Calibri', size=14, bold=True)
HEADER_FONT = Font(name='Calibri', size=12, bold=True)
NORMAL_FONT = Font(name='Calibri', size=11)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
HEADER_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

def read_excel_data(file_path):
    """
    Read data from an Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        DataFrame: The data read from the Excel file
    """
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            return None
        
        # Read the Excel file
        df = pd.read_excel(file_path)
        
        # Validate required columns
        required_columns = ['date', 'amount', 'category']
        if not all(col in df.columns for col in required_columns):
            missing = [col for col in required_columns if col not in df.columns]
            logger.error(f"Missing required columns: {missing}")
            return None
        
        # Ensure date column is datetime
        df['date'] = pd.to_datetime(df['date'])
        
        # Sort by date
        df = df.sort_values('date')
        
        return df
    
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}")
        return None

def generate_category_pie_chart(df, output_path='category_pie.png'):
    """
    Generate a pie chart of expenses by category.
    
    Args:
        df: DataFrame containing the expense data
        output_path: Path to save the chart image
    
    Returns:
        str: Path to the saved chart image
    """
    try:
        # Group by category and sum amounts
        category_sums = df.groupby('category')['amount'].sum()
        
        # Create the figure and axis
        plt.figure(figsize=(8, 6))
        
        # Generate pie chart
        wedges, texts, autotexts = plt.pie(
            category_sums,
            labels=None,
            autopct='%1.1f%%',
            startangle=90,
            shadow=False,
            explode=[0.05] * len(category_sums),
            textprops={'fontsize': 10}
        )
        
        # Customize appearance
        plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
        plt.title('Expenses by Category', fontsize=14, fontweight='bold', pad=20)
        
        # Add legend
        plt.legend(
            wedges,
            category_sums.index,
            title="Categories",
            loc="center left",
            bbox_to_anchor=(1, 0, 0.5, 1)
        )
        
        # Tighten the layout
        plt.tight_layout()
        
        # Save the chart
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    except Exception as e:
        logger.error(f"Error generating pie chart: {str(e)}")
        return None

def generate_time_series_chart(df, output_path='time_series.png'):
    """
    Generate a time series line chart of expenses over time.
    
    Args:
        df: DataFrame containing the expense data
        output_path: Path to save the chart image
    
    Returns:
        str: Path to the saved chart image
    """
    try:
        # Group by date and sum amounts
        daily_expenses = df.groupby('date')['amount'].sum()
        
        # Create the figure and axis
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Generate line chart
        ax.plot(
            daily_expenses.index, 
            daily_expenses.values,
            marker='o',
            linestyle='-',
            linewidth=2,
            markersize=6,
            color='#4285F4'
        )
        
        # Customize appearance
        ax.set_xlabel('Date', fontsize=12)
        ax.set_ylabel('Total Expense Amount', fontsize=12)
        ax.set_title('Daily Expenses Over Time', fontsize=14, fontweight='bold')
        
        # Format x-axis to show dates nicely
        date_format = mdates.DateFormatter('%Y-%m-%d')
        ax.xaxis.set_major_formatter(date_format)
        fig.autofmt_xdate()  # Rotate date labels
        
        # Add grid
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # Add text labels for each point
        for x, y in zip(daily_expenses.index, daily_expenses.values):
            ax.annotate(
                f'${y:.2f}',
                (x, y),
                textcoords="offset points",
                xytext=(0, 10),
                ha='center',
                fontsize=9
            )
        
        # Tighten the layout
        plt.tight_layout()
        
        # Save the chart
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    except Exception as e:
        logger.error(f"Error generating time series chart: {str(e)}")
        return None

def generate_payment_method_bar_chart(df, output_path='payment_method_bar.png'):
    """
    Generate a bar chart of expenses by payment method.
    
    Args:
        df: DataFrame containing the expense data
        output_path: Path to save the chart image
    
    Returns:
        str: Path to the saved chart image
    """
    try:
        # Check if payment_method column exists
        if 'payment_method' not in df.columns:
            logger.warning("Payment method column not found, skipping bar chart")
            return None
        
        # Filter out rows with no payment method
        df_filtered = df.dropna(subset=['payment_method'])
        
        # Group by payment method and sum amounts
        payment_sums = df_filtered.groupby('payment_method')['amount'].sum().sort_values(ascending=False)
        
        # Create the figure and axis
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Generate bar chart
        bars = ax.bar(
            payment_sums.index,
            payment_sums.values,
            color='#34A853',
            width=0.6,
            edgecolor='black',
            linewidth=0.5
        )
        
        # Customize appearance
        ax.set_xlabel('Payment Method', fontsize=12)
        ax.set_ylabel('Total Amount', fontsize=12)
        ax.set_title('Expenses by Payment Method', fontsize=14, fontweight='bold')
        
        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width()/2.,
                height + 5,
                f'${height:.2f}',
                ha='center',
                va='bottom',
                fontsize=10
            )
        
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45, ha='right')
        
        # Add grid lines
        ax.grid(True, linestyle='--', alpha=0.3, axis='y')
        
        # Adjust layout
        plt.tight_layout()
        
        # Save the chart
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    except Exception as e:
        logger.error(f"Error generating payment method bar chart: {str(e)}")
        return None

def generate_merchant_bar_chart(df, output_path='merchant_bar.png', top_n=5):
    """
    Generate a bar chart of expenses by top merchants.
    
    Args:
        df: DataFrame containing the expense data
        output_path: Path to save the chart image
        top_n: Number of top merchants to include
        
    Returns:
        str: Path to the saved chart image
    """
    try:
        # Check if merchant column exists
        if 'merchant' not in df.columns:
            logger.warning("Merchant column not found, skipping bar chart")
            return None
        
        # Filter out rows with no merchant
        df_filtered = df.dropna(subset=['merchant'])
        
        # Group by merchant and sum amounts
        merchant_sums = df_filtered.groupby('merchant')['amount'].sum().sort_values(ascending=False)
        
        # Take top N merchants
        top_merchants = merchant_sums.head(top_n)
        
        # Create the figure and axis
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Generate horizontal bar chart
        bars = ax.barh(
            top_merchants.index,
            top_merchants.values,
            color='#FBBC05',
            height=0.6,
            edgecolor='black',
            linewidth=0.5
        )
        
        # Customize appearance
        ax.set_xlabel('Total Amount', fontsize=12)
        ax.set_ylabel('Merchant', fontsize=12)
        ax.set_title(f'Top {top_n} Merchants by Expense Amount', fontsize=14, fontweight='bold')
        
        # Add value labels
        for bar in bars:
            width = bar.get_width()
            ax.text(
                width + 5,
                bar.get_y() + bar.get_height()/2.,
                f'${width:.2f}',
                ha='left',
                va='center',
                fontsize=10
            )
        
        # Add grid lines
        ax.grid(True, linestyle='--', alpha=0.3, axis='x')
        
        # Adjust layout
        plt.tight_layout()
        
        # Save the chart
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    except Exception as e:
        logger.error(f"Error generating merchant bar chart: {str(e)}")
        return None

def generate_category_trend_chart(df, output_path='category_trend.png', top_n=3):
    """
    Generate a line chart showing trends of top categories over time.
    
    Args:
        df: DataFrame containing the expense data
        output_path: Path to save the chart image
        top_n: Number of top categories to include
        
    Returns:
        str: Path to the saved chart image
    """
    try:
        # Find top N categories by total amount
        top_categories = df.groupby('category')['amount'].sum().sort_values(ascending=False).head(top_n).index
        
        # Group by date and category
        df_trend = df[df['category'].isin(top_categories)].copy()
        df_pivot = df_trend.pivot_table(
            index='date',
            columns='category',
            values='amount',
            aggfunc='sum',
            fill_value=0
        )
        
        # Create the figure and axis
        fig, ax = plt.subplots(figsize=(12, 6))
        
        # Plot each category
        for category in top_categories:
            if category in df_pivot.columns:
                ax.plot(
                    df_pivot.index,
                    df_pivot[category],
                    marker='o',
                    label=category,
                    linewidth=2,
                    markersize=5
                )
        
        # Customize appearance
        ax.set_xlabel('Date', fontsize=12)
        ax.set_ylabel('Amount', fontsize=12)
        ax.set_title(f'Expense Trends for Top {top_n} Categories', fontsize=14, fontweight='bold')
        
        # Format x-axis to show dates nicely
        date_format = mdates.DateFormatter('%Y-%m-%d')
        ax.xaxis.set_major_formatter(date_format)
        fig.autofmt_xdate()  # Rotate date labels
        
        # Add legend
        ax.legend(loc='upper left', frameon=True, fancybox=True, shadow=True)
        
        # Add grid
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # Adjust layout
        plt.tight_layout()
        
        # Save the chart
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()
        
        return output_path
    
    except Exception as e:
        logger.error(f"Error generating category trend chart: {str(e)}")
        return None

def create_summary_table(df):
    """
    Create a summary table of expenses by category.
    
    Args:
        df: DataFrame containing the expense data
        
    Returns:
        pandas.DataFrame: Summary table with categories, amounts, and percentages
    """
    try:
        # Group by category
        category_summary = df.groupby('category')['amount'].agg(['sum', 'count'])
        category_summary.columns = ['Total Amount', 'Count']
        
        # Calculate percentages
        category_summary['Percentage'] = category_summary['Total Amount'] / category_summary['Total Amount'].sum() * 100
        
        # Sort by total amount descending
        category_summary = category_summary.sort_values('Total Amount', ascending=False)
        
        # Format values
        category_summary['Total Amount'] = category_summary['Total Amount'].map('${:,.2f}'.format)
        category_summary['Percentage'] = category_summary['Percentage'].map('{:,.1f}%'.format)
        
        return category_summary
    
    except Exception as e:
        logger.error(f"Error creating summary table: {str(e)}")
        return None

def add_charts_to_excel(file_path, chart_paths, summary_table):
    """
    Add generated charts to the Excel file.
    
    Args:
        file_path: Path to the Excel file
        chart_paths: List of paths to chart images
        summary_table: Summary table to add to the file
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Create a new sheet for the visualizations if it doesn't exist
        if 'Visualizations' in workbook.sheetnames:
            sheet = workbook['Visualizations']
        else:
            sheet = workbook.create_sheet('Visualizations')
        
        # Clear any existing content
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
        
        # Add title to the sheet
        sheet['A1'] = 'Expense Analysis & Visualizations'
        sheet['A1'].font = Font(name='Calibri', size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('A1:I1')
        
        # Add date of analysis
        sheet['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        sheet['A2'].font = Font(name='Calibri', size=10, italic=True)
        sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('A2:I2')
        
        # Set column widths
        for col in range(1, 10):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 15
        
        # Add summary statistics
        sheet['A4'] = 'Summary Statistics'
        sheet['A4'].font = TITLE_FONT
        sheet.merge_cells('A4:D4')
        
        # Add the summary table
        if summary_table is not None:
            # Add headers
            headers = ['Category', 'Total Amount', 'Count', 'Percentage']
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=5, column=col)
                cell.value = header
                cell.font = HEADER_FONT
                cell.alignment = ALIGN_CENTER
                cell.fill = HEADER_FILL
                cell.border = BORDER
            
            # Add data
            for row_idx, (category, data) in enumerate(summary_table.iterrows(), start=6):
                # Category
                cell = sheet.cell(row=row_idx, column=1)
                cell.value = category
                cell.font = NORMAL_FONT
                cell.border = BORDER
                
                # Amount
                cell = sheet.cell(row=row_idx, column=2)
                cell.value = data['Total Amount']
                cell.font = NORMAL_FONT
                cell.alignment = Alignment(horizontal='right')
                cell.border = BORDER
                
                # Count
                cell = sheet.cell(row=row_idx, column=3)
                cell.value = data['Count']
                cell.font = NORMAL_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = BORDER
                
                # Percentage
                cell = sheet.cell(row=row_idx, column=4)
                cell.value = data['Percentage']
                cell.font = NORMAL_FONT
                cell.alignment = Alignment(horizontal='right')
                cell.border = BORDER
        
        # Add charts
        row_position = 5 + len(summary_table) + 3 if summary_table is not None else 7
        
        # Add title for charts section
        sheet.cell(row=row_position, column=1).value = 'Expense Visualizations'
        sheet.cell(row=row_position, column=1).font = TITLE_FONT
        sheet.merge_cells(f'A{row_position}:D{row_position}')
        row_position += 1
        
        # Add each chart
        chart_row = row_position
        chart_col = 1
        
        for i, chart_path in enumerate(chart_paths):
            if chart_path and os.path.exists(chart_path):
                try:
                    # Add the chart with appropriate sizing
                    img = Image(chart_path)
                    # Scale the image to fit in Excel cells
                    img.width = 400
                    img.height = 300
                    
                    # Position the image
                    sheet.add_image(img, f'{get_column_letter(chart_col)}{chart_row}')
                    
                    # Move to next position
                    if chart_col == 1:
                        chart_col = 6  # Move to right side
                    else:
                        chart_col = 1  # Reset to left
                        chart_row += 20  # Move down for next row of charts
                except Exception as e:
                    logger.error(f"Error adding chart {chart_path}: {str(e)}")
        
        # Save the workbook with a new name
        output_name = os.path.splitext(file_path)[0] + '_analyzed.xlsx'
        workbook.save(output_name)
        logger.info(f"Analysis saved to {output_name}")
        
        return True
    
    except Exception as e:
        logger.error(f"Error adding charts to Excel: {str(e)}")
        return False

def analyze_excel_file(file_path):
    """
    Analyze an Excel file and create visualizations.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        str: Path to the analyzed Excel file, or None if analysis failed
    """
    try:
        # Read the data
        df = read_excel_data(file_path)
        if df is None:
            return None
        
        # Create a folder for temporary chart images
        temp_dir = 'temp_charts'
        os.makedirs(temp_dir, exist_ok=True)
        
        # Generate charts
        chart_paths = []
        
        # Pie chart of categories
        pie_chart_path = os.path.join(temp_dir, 'category_pie.png')
        chart_paths.append(generate_category_pie_chart(df, pie_chart_path))
        
        # Time series chart
        time_series_path = os.path.join(temp_dir, 'time_series.png')
        chart_paths.append(generate_time_series_chart(df, time_series_path))
        
        # Payment method bar chart
        payment_chart_path = os.path.join(temp_dir, 'payment_method.png')
        chart_paths.append(generate_payment_method_bar_chart(df, payment_chart_path))
        
        # Merchant bar chart
        merchant_chart_path = os.path.join(temp_dir, 'merchant.png')
        chart_paths.append(generate_merchant_bar_chart(df, merchant_chart_path))
        
        # Category trend chart
        trend_chart_path = os.path.join(temp_dir, 'category_trend.png')
        chart_paths.append(generate_category_trend_chart(df, trend_chart_path))
        
        # Create summary table
        summary_table = create_summary_table(df)
        
        # Add charts to Excel
        output_name = os.path.splitext(file_path)[0] + '_analyzed.xlsx'
        if add_charts_to_excel(file_path, chart_paths, summary_table):
            logger.info(f"Successfully analyzed Excel file: {file_path}")
            return output_name
        else:
            logger.error(f"Failed to add charts to Excel file: {file_path}")
            return None
        
    except Exception as e:
        logger.error(f"Error analyzing Excel file: {str(e)}")
        return None

if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python excel_visualizer.py <path_to_excel_file>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    
    output_file = analyze_excel_file(file_path)
    if output_file:
        print(f"Analysis completed. Output saved to: {output_file}")
    else:
        print("Analysis failed. Check the logs for details.")
        sys.exit(1)